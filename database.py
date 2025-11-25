import psycopg2
import os
from dotenv import load_dotenv
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill

load_dotenv()

DB_PARAMS = {
    "dbname": os.getenv("PG_DB"),
    "user": os.getenv("PG_USER"),
    "password": os.getenv("PG_PASSWORD"),
    "host": os.getenv("PG_HOST"),
    "port": os.getenv("PG_PORT"),
}


def connect():
    return psycopg2.connect(**DB_PARAMS)


def init_db():
    """Инициализация базы данных. Таблица создается только если её нет."""
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS channels (
                    id BIGINT PRIMARY KEY,
                    title TEXT,
                    members INTEGER,
                    videos INTEGER DEFAULT 0,
                    date_added TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    type TEXT DEFAULT 'unknown',
                    link TEXT DEFAULT ''
                );
            """)
        conn.commit()


def add_or_update_channel(chat_id, title, members, chat_type="unknown", link=""):
    """Добавляет новый канал или обновляет существующий по id."""
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                INSERT INTO channels (id, title, members, type, link)
                VALUES (%s, %s, %s, %s, %s)
                ON CONFLICT (id) DO UPDATE
                SET title = EXCLUDED.title,
                    members = EXCLUDED.members,
                    type = EXCLUDED.type,
                    link = EXCLUDED.link;
            """, (chat_id, title, members, chat_type, link))
        conn.commit()


def update_channel_status(chat_id, title=None, members=None, chat_type=None, link=None):
    """Обновляет данные канала по id."""
    with connect() as conn:
        with conn.cursor() as cur:
            updates = []
            values = []

            if title is not None:
                updates.append("title = %s")
                values.append(title)
            if members is not None:
                updates.append("members = %s")
                values.append(members)
            if chat_type is not None:
                updates.append("type = %s")
                values.append(chat_type)
            if link is not None:
                updates.append("link = %s")
                values.append(link)

            if not updates:
                return

            values.append(chat_id)
            query = f"UPDATE channels SET {', '.join(updates)} WHERE id = %s;"
            cur.execute(query, values)
        conn.commit()


def increment_video_count(chat_id):
    """Увеличивает счетчик видео на 1."""
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                UPDATE channels
                SET videos = videos + 1
                WHERE id = %s;
            """, (chat_id,))
        conn.commit()


def get_channels():
    """Возвращает все каналы/группы."""
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                SELECT id, title, members, videos, date_added, type, link
                FROM channels
                ORDER BY date_added;
            """)
            return cur.fetchall()


def delete_channel(chat_id):
    """Удаляет конкретный канал по id (для функции 'Покинуть все чаты')."""
    with connect() as conn:
        with conn.cursor() as cur:
            cur.execute("DELETE FROM channels WHERE id = %s;", (chat_id,))
        conn.commit()


def export_excel():
    """Экспорт всей таблицы в Excel."""
    channels = get_channels()
    if not channels:
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Каналы и группы"
    ws.append(["ID", "Название", "Участники", "Отправлено видео", "Дата добавления", "Тип", "Ссылка"])

    red_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")
    green_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")

    for row in channels:
        id_, title, members, videos, date_added, chat_type, link = row
        date_str = date_added.strftime('%Y-%m-%d %H:%M') if isinstance(date_added, datetime) else str(date_added)
        data = [id_, title, members, videos, date_str, chat_type, link]
        ws.append(data)
        fill = red_fill if chat_type in ["left", "kicked"] else green_fill
        for col in range(1, len(data)+1):
            ws.cell(row=ws.max_row, column=col).fill = fill

    file_path = "channels_export.xlsx"
    wb.save(file_path)
    return file_path
